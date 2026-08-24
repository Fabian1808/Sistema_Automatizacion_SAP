import os
import threading
from pathlib import Path

from PySide6.QtCore import Signal, QTimer
from PySide6.QtWidgets import (
    QCheckBox,
    QDialog,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QPlainTextEdit,
    QProgressBar,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from ..core.error_catalog import explain_error
from ..sap.sap_connection import SapConnection
from ..services.report_service import ReportService
from ..services.state_service import DocumentState, StateService
from ..utils.ids import analyze_ids


class RunPanel(QWidget):
    run_requested = Signal(list, bool, str, bool)
    cancel_requested = Signal()
    _sap_check_done = Signal(bool, str)

    def __init__(self, config, doc_label, report_name):
        super().__init__()
        self._config = config
        self._doc_label = doc_label
        self._report_name = report_name
        self._valid_ids = []
        self._last_ids = []
        self._last_dry = False
        self._failed_ids = []
        self._running = False
        self._current_batch_id = None
        self._state_service = StateService()
        self._errors_by_doc = {}
        self._ok_count = 0
        self._fail_count = 0
        self._worker = None
        self._sap_check_done.connect(self._on_sap_check)
        self._resume_timer = QTimer(self)
        self._resume_timer.setSingleShot(True)
        self._resume_timer.setInterval(300)
        self._resume_timer.timeout.connect(self._check_resume_available)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        buttons = QHBoxLayout()
        self.btn_excel = QPushButton("Importar Excel")
        self.btn_csv = QPushButton("Importar CSV")
        self.btn_excel.clicked.connect(self._import_file)
        self.btn_csv.clicked.connect(self._import_file)
        buttons.addWidget(self.btn_excel)
        buttons.addWidget(self.btn_csv)
        buttons.addStretch(1)
        layout.addLayout(buttons)

        self.text_edit = QPlainTextEdit()
        self.text_edit.setPlaceholderText(
            "Pegue los números de documento aquí, uno por línea."
        )
        self.text_edit.textChanged.connect(self._update_counts)
        layout.addWidget(self.text_edit, 2)

        self.counts_label = QLabel("0 documentos válidos")
        layout.addWidget(self.counts_label)

        self.dry_run_check = QCheckBox(
            "Modo prueba (solo navegación, no imprime ni modifica datos)"
        )
        layout.addWidget(self.dry_run_check)

        self.btn_process = QPushButton(f"PROCESAR {self._doc_label.upper()}")
        self.btn_process.setObjectName("primary")
        self.btn_process.setEnabled(False)
        self.btn_process.clicked.connect(self._on_process)
        layout.addWidget(self.btn_process)

        self.btn_resume = QPushButton("Reanudar lote anterior")
        self.btn_resume.setVisible(False)
        self.btn_resume.clicked.connect(self._on_resume)
        layout.addWidget(self.btn_resume)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress_label = QLabel("")
        self.progress_label.setObjectName("subtitle")
        self.progress_label.setVisible(False)
        layout.addWidget(self.progress_label)
        layout.addWidget(self.progress)

        self.counters_label = QLabel("")
        self.counters_label.setObjectName("subtitle")
        self.counters_label.setVisible(False)
        layout.addWidget(self.counters_label)

        self.btn_cancel = QPushButton("Cancelar proceso")
        self.btn_cancel.setObjectName("dangerButton")
        self.btn_cancel.setVisible(False)
        self.btn_cancel.clicked.connect(self._on_cancel)
        layout.addWidget(self.btn_cancel)

        self.results_table = QTableWidget(0, 3)
        self.results_table.setHorizontalHeaderLabels(["Documento", "Estado", "Archivo"])
        self.results_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Stretch
        )
        self.results_table.verticalHeader().setVisible(False)
        self.results_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.results_table.cellDoubleClicked.connect(self._on_row_detail)
        layout.addWidget(self.results_table, 3)

        self.btn_retry = QPushButton("Reintentar errores")
        self.btn_retry.setVisible(False)
        self.btn_retry.clicked.connect(self._on_retry)
        layout.addWidget(self.btn_retry)

    def _import_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Importar archivo", "", "Excel o CSV (*.xlsx *.xls *.csv)"
        )
        if not path:
            return
        text = self._read_file(path)
        if text is None:
            return
        self.text_edit.setPlainText(text)

    def _read_file(self, path):
        if path.lower().endswith((".xlsx", ".xls")):
            try:
                from openpyxl import load_workbook
            except ImportError:
                QMessageBox.critical(
                    self, "Error", "La biblioteca openpyxl no está instalada."
                )
                return None
            try:
                workbook = load_workbook(path, read_only=True, data_only=True)
                sheet = workbook.active
                lines = []
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            lines.append(str(cell.value).strip())
                return "\n".join(lines)
            except Exception as exc:
                QMessageBox.critical(
                    self,
                    "Error al leer el archivo",
                    f"No se pudo leer el archivo:\n{exc}",
                )
                return None
        try:
            with open(path, "r", encoding="utf-8-sig", errors="replace") as fh:
                return fh.read()
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Error al leer el archivo",
                f"No se pudo leer el archivo:\n{exc}",
            )
            return None

    def _update_counts(self):
        result = analyze_ids(self.text_edit.toPlainText().splitlines())
        self._valid_ids = result["valid"]
        parts = [f"{len(result['valid'])} {self._doc_label}s vǭlidos"]
        if result["duplicates"]:
            parts.append(f"{len(result['duplicates'])} duplicados")
        if result["invalid"]:
            parts.append(f"{len(result['invalid'])} invǭlidos")
        self.counts_label.setText(" | ".join(parts))

        # Consulta SQLite diferida: no se ejecuta en cada tecla
        self._resume_timer.start()
        self.btn_process.setEnabled(bool(result["valid"]) and not self._running)

    def _check_resume_available(self):
        if not self._valid_ids:
            self.btn_resume.setVisible(False)
            return
        # Buscar lotes recientes del mismo módulo
        batches = self._state_service.list_batches(limit=10)
        for batch in batches:
            if batch.get("module_id") != self._doc_label.lower():
                continue
            pending = self._state_service.get_documents(
                batch["batch_id"],
                [DocumentState.PENDING, DocumentState.RETRY, DocumentState.FAILED],
            )
            if pending:
                self._last_batch_id = batch["batch_id"]
                self.btn_resume.setVisible(True)
                self.btn_resume.setText(f"Reanudar lote ({len(pending)} pendientes)")
                return
        self.btn_resume.setVisible(False)

    def _on_process(self):
        if self._running or not self._valid_ids:
            return
        self.btn_process.setEnabled(False)
        self.progress_label.setVisible(True)
        self.progress_label.setText("Verificando conexi��n con SAP...")
        timeout = int(self._config.get("sap_timeout_seconds") or 30)
        threading.Thread(
            target=self._sap_check_worker, args=(timeout,), daemon=True
        ).start()

    def _sap_check_worker(self, timeout: int):
        try:
            ok = SapConnection(timeout=timeout).is_running()
            message = "" if ok else "SAP GUI no estǭ abierto."
        except Exception as exc:
            # El chequeo es preventivo: si falla, no bloquea el proceso
            ok, message = True, ""
        self._sap_check_done.emit(ok, message)

    def _on_sap_check(self, ok: bool, message: str):
        if self._running:
            return
        if not ok:
            self.progress_label.setVisible(False)
            self.btn_process.setEnabled(bool(self._valid_ids))
            QMessageBox.warning(
                self,
                "SAP no detectado",
                f"{message} Abra SAP GUI, inicie sesi��n y vuelva a intentar.",
            )
            return
        self.progress_label.setVisible(False)
        self.btn_process.setEnabled(bool(self._valid_ids))
        dry_run = self.dry_run_check.isChecked()
        if self._config.get("confirm_before_process"):
            mode = (
                "en modo prueba (sin imprimir)"
                if dry_run
                else "utilizando la sesi��n SAP actualmente abierta"
            )
            answer = QMessageBox.question(
                self,
                "Confirmar proceso",
                f"Se procesarǭn {len(self._valid_ids)} {self._doc_label}s {mode}.\n\n"
                "Verifique que SAP corresponda al entorno correcto.\n\n"
                "��Desea continuar?",
            )
            if answer != QMessageBox.StandardButton.Yes:
                return
        self.run_requested.emit(self._valid_ids, dry_run, "", False)

    def _on_resume(self):
        if not hasattr(self, "_last_batch_id") or self._running:
            return
        dry_run = self.dry_run_check.isChecked()
        answer = QMessageBox.question(
            self,
            "Reanudar lote",
            f"Reanudar lote anterior con {len(self._valid_ids)} documentos?\n"
            "Se saltarán los ya procesados y se reintentarán los fallidos.",
        )
        if answer != QMessageBox.StandardButton.Yes:
            return
        self.run_requested.emit(self._valid_ids, dry_run, self._last_batch_id, True)

    def start_run(self, ids, batch_id=""):
        self._running = True
        self._current_batch_id = batch_id
        self._last_ids = list(ids)
        self._last_dry = self.dry_run_check.isChecked()
        self._failed_ids = []
        self._errors_by_doc = {}
        self._ok_count = 0
        self._fail_count = 0
        self.results_table.setRowCount(0)
        self.progress.setVisible(True)
        self.progress_label.setVisible(True)
        self.counters_label.setVisible(True)
        self.counters_label.setText(f"⏳ Pendientes: {len(ids)}")
        self.progress.setValue(0)
        self.progress_label.setText("Preparando...")
        self.btn_process.setEnabled(False)
        self.btn_resume.setVisible(False)
        self.btn_retry.setVisible(False)
        self.btn_cancel.setVisible(True)
        self.text_edit.setReadOnly(True)
        self.btn_excel.setEnabled(False)
        self.btn_csv.setEnabled(False)
        self.dry_run_check.setEnabled(False)

    def on_progress(self, done, total, current):
        if not self._running:
            return
        self.progress.setMaximum(total)
        self.progress.setValue(done)
        pending = total - done - (1 if done < total else 0)
        self.progress_label.setText(
            f"Procesando {self._doc_label} {min(done + 1, total)} de {total}..."
            if done < total
            else "Finalizando..."
        )
        self.counters_label.setText(
            f"✅ {self._ok_count}   ❌ {self._fail_count}   ⏳ {max(pending, 0)}"
        )

    def on_document(self, result):
        row = self.results_table.rowCount()
        self.results_table.insertRow(row)
        self.results_table.setItem(row, 0, QTableWidgetItem(result.document_id))
        self.results_table.setItem(
            row, 1, QTableWidgetItem("OK" if result.ok else "ERROR")
        )
        self.results_table.setItem(
            row, 2, QTableWidgetItem(result.file_path if result.file_path else "-")
        )
        if not result.ok:
            self._fail_count += 1
            self._errors_by_doc[result.document_id] = result.error
            for col in range(3):
                item = self.results_table.item(row, col)
                if item:
                    item.setToolTip(result.error)
        else:
            self._ok_count += 1
        self.results_table.scrollToBottom()

    def _on_cancel(self):
        if self._worker is not None:
            self.btn_cancel.setEnabled(False)
            self.progress_label.setText("Cancelando tras el documento en curso...")
            self.cancel_requested.emit()

    def _on_row_detail(self, row: int, column: int):
        doc_item = self.results_table.item(row, 0)
        state_item = self.results_table.item(row, 1)
        if doc_item is None or state_item is None:
            return
        if state_item.text() != "ERROR":
            return
        self._show_error_detail(doc_item.text())

    def _show_error_detail(self, document_id: str):
        info = explain_error(self._errors_by_doc.get(document_id, ""))
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Detalle del error — {document_id}")
        layout = QVBoxLayout(dialog)
        header = QLabel(f"<b>Documento:</b> {document_id}")
        motivo = QLabel(f"<b>Motivo:</b> {info['motivo']}")
        detalle = QLabel(f"<b>Detalle:</b> {info['detalle']}")
        detalle.setWordWrap(True)
        etapa = QLabel(f"<b>Etapa:</b> {info['etapa']}")
        reco = QLabel(f"<b>Recomendaci��n:</b> {info['recomendacion']}")
        reco.setWordWrap(True)
        for w in (header, motivo, detalle, etapa, reco):
            layout.addWidget(w)
        close = QPushButton("Cerrar")
        close.clicked.connect(dialog.accept)
        layout.addWidget(close)
        dialog.resize(460, 240)
        dialog.exec()

    def on_finished(self, results):
        self._running = False
        self._worker = None
        self.text_edit.setReadOnly(False)
        self.btn_excel.setEnabled(True)
        self.btn_csv.setEnabled(True)
        self.dry_run_check.setEnabled(True)
        self.btn_process.setEnabled(bool(self._valid_ids))
        self.btn_cancel.setVisible(False)
        self.btn_cancel.setEnabled(True)
        ok_count = sum(1 for r in results if r.ok)
        err_count = len(results) - ok_count
        cancelled = getattr(self, "_was_cancelled", False)
        self._was_cancelled = False
        if cancelled:
            estado = "CANCELADO POR EL USUARIO"
            self.progress_label.setText("Proceso cancelado")
        elif err_count == 0:
            estado = "COMPLETADO"
            self.progress_label.setText("Proceso finalizado")
        else:
            estado = "FINALIZADO CON ERRORES"
            self.progress_label.setText("Proceso finalizado con errores")
        self.counters_label.setText(
            f"✅ {ok_count}   ❌ {err_count}"
        )
        self._failed_ids = [r.document_id for r in results if not r.ok]
        self.btn_retry.setVisible(bool(self._failed_ids) and not cancelled)
        mode = " (modo prueba)" if self._last_dry else ""
        box = QMessageBox(self)
        box.setWindowTitle("Resumen")
        box.setText(
            f"{estado}{mode}\n\n"
            f"Total procesado: {len(results)}\n"
            f"Correctas: {ok_count}\n"
            f"Errores: {err_count}"
        )
        btn_folder = box.addButton("Abrir carpeta", QMessageBox.ButtonRole.ActionRole)
        btn_export = box.addButton("Exportar reporte", QMessageBox.ButtonRole.ActionRole)
        box.addButton("Cerrar", QMessageBox.ButtonRole.AcceptRole)
        box.exec()
        clicked = box.clickedButton()
        if clicked == btn_folder:
            folder = Path(self._config.output_folder)
            folder.mkdir(parents=True, exist_ok=True)
            os.startfile(str(folder))
        elif clicked == btn_export:
            report = ReportService().export(
                results, self._report_name, self._config.output_folder
            )
            QMessageBox.information(
                self, "Reporte", f"Reporte generado:\n{report}"
            )

    def _on_retry(self):
        if self._running or not self._failed_ids:
            return
        self.run_requested.emit(self._failed_ids, self._last_dry, "", False)
